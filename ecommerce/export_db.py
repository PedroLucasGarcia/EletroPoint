import os
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ecommerce.settings')
django.setup()

import openpyxl
from openpyxl.styles import Font
from store.models import OrderItem, Order, Product, Customer, Category, Brand, ShippingAddress

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove a aba padrão "Sheet"

HEADER_FONT = Font(bold=True)


def write_sheet(name, headers, rows):
    ws = wb.create_sheet(title=name)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
    for row in rows:
        ws.append(row)
    # ajusta largura das colunas
    for col_idx, header in enumerate(headers, start=1):
        max_len = max([len(str(header))] + [len(str(r[col_idx - 1])) for r in rows]) if rows else len(str(header))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)


# ── ORDERITEMS ───────────────────────────────────────────
headers = ['id', 'order_id', 'product_id', 'product', 'category', 'brand', 'quantity',
           'color', 'storage', 'custom_price', 'total', 'date_added', 'complete', 'customer']
rows = []
for item in OrderItem.objects.select_related('product', 'order', 'order__customer', 'product__category', 'product__brand'):
    rows.append([
        item.id,
        item.order.id if item.order else '',
        item.product.id if item.product else '',
        item.product.name if item.product else '',
        item.product.category.name if item.product and item.product.category else '',
        item.product.brand.name if item.product and item.product.brand else '',
        item.quantity,
        item.color,
        item.storage,
        item.custom_price,
        float(item.get_total),
        item.date_added.replace(tzinfo=None) if item.date_added else '',
        item.order.complete if item.order else '',
        item.order.customer.name if item.order and item.order.customer else '',
    ])
write_sheet('OrderItems', headers, rows)


# ── PRODUCTS ─────────────────────────────────────────────
headers = ['id', 'name', 'category', 'brand', 'price', 'slug']
rows = []
for p in Product.objects.select_related('category', 'brand'):
    rows.append([
        p.id,
        p.name,
        p.category.name if p.category else '',
        p.brand.name if p.brand else '',
        float(p.price),
        p.slug,
    ])
write_sheet('Products', headers, rows)


# ── ORDERS ───────────────────────────────────────────────
headers = ['id', 'customer_id', 'customer', 'date_ordered', 'complete', 'transaction_id', 'total', 'items_count']
rows = []
for o in Order.objects.select_related('customer'):
    rows.append([
        o.id,
        o.customer.id if o.customer else '',
        o.customer.name if o.customer else '',
        o.date_ordered.replace(tzinfo=None) if o.date_ordered else '',
        o.complete,
        o.transaction_id,
        float(o.get_cart_total),
        o.get_cart_items,
    ])
write_sheet('Orders', headers, rows)


# ── CUSTOMERS ────────────────────────────────────────────
headers = ['id', 'name', 'email', 'user_id']
rows = []
for c in Customer.objects.all():
    rows.append([
        c.id,
        c.name,
        c.email,
        c.user.id if c.user else '',
    ])
write_sheet('Customers', headers, rows)


# ── CATEGORIES ───────────────────────────────────────────
headers = ['id', 'name', 'slug']
rows = [[c.id, c.name, c.slug] for c in Category.objects.all()]
write_sheet('Categories', headers, rows)


# ── BRANDS ───────────────────────────────────────────────
headers = ['id', 'name', 'slug']
rows = [[b.id, b.name, b.slug] for b in Brand.objects.all()]
write_sheet('Brands', headers, rows)


# ── SHIPPING ADDRESSES ───────────────────────────────────
headers = ['id', 'customer_id', 'customer', 'order_id', 'address', 'zipcode', 'city', 'country', 'date_added']
rows = []
for s in ShippingAddress.objects.select_related('customer', 'order'):
    rows.append([
        s.id,
        s.customer.id if s.customer else '',
        s.customer.name if s.customer else '',
        s.order.id if s.order else '',
        s.address,
        s.zipcode,
        s.city,
        s.country,
        s.date_added.replace(tzinfo=None) if s.date_added else '',
    ])
write_sheet('ShippingAddresses', headers, rows)


wb.save('EletroPoint_DB.xlsx')
print("Exportação concluída: EletroPoint_DB.xlsx")
print("Abas criadas:", wb.sheetnames)