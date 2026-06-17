import os
import django
from datetime import date, timedelta

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ecommerce.settings')
django.setup()

import openpyxl
from openpyxl.styles import Font
from store.models import Order

HEADER_FONT = Font(bold=True)

# ── Determina o intervalo de datas com base nas encomendas existentes ──
orders = Order.objects.exclude(date_ordered__isnull=True)
if orders.exists():
    min_date = min(o.date_ordered.date() for o in orders)
    max_date = max(o.date_ordered.date() for o in orders)
else:
    min_date = date.today().replace(month=1, day=1)
    max_date = date.today()

# Margem de segurança (início do ano até hoje + 1 ano)
start_date = date(min_date.year, 1, 1)
end_date = date(max_date.year + 1, 12, 31)

MESES_PT = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
            'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
DIAS_PT = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']

headers = ['Date', 'Year', 'Month', 'MonthName', 'MonthShort', 'Quarter',
           'Day', 'Weekday', 'WeekdayName', 'YearMonth', 'IsWeekend']

rows = []
current = start_date
while current <= end_date:
    rows.append([
        current,
        current.year,
        current.month,
        MESES_PT[current.month - 1],
        MESES_PT[current.month - 1][:3],
        f"T{(current.month - 1) // 3 + 1}",
        current.day,
        current.isoweekday(),  # 1=Segunda ... 7=Domingo
        DIAS_PT[current.isoweekday() - 1],
        f"{current.year}-{current.month:02d}",
        1 if current.isoweekday() >= 6 else 0,
    ])
    current += timedelta(days=1)

# ── Abre o ficheiro existente e adiciona a aba DateTable ──
wb = openpyxl.load_workbook('EletroPoint_DB.xlsx')

if 'DateTable' in wb.sheetnames:
    del wb['DateTable']

ws = wb.create_sheet(title='DateTable')
ws.append(headers)
for cell in ws[1]:
    cell.font = HEADER_FONT
for row in rows:
    ws.append(row)

# largura das colunas
for col_idx, header in enumerate(headers, start=1):
    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(header) + 2, 12)

wb.save('EletroPoint_DB.xlsx')
print(f"DateTable adicionada: {start_date} até {end_date} ({len(rows)} dias)")